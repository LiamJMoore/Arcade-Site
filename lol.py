#!/usr/bin/env python3
“””
MULTI-WEEK PLANNED JOBS ANALYSER v3.0 — LEGENDARY EDITION
══════════════════════════════════════════════════════════
A professional-grade desktop analytics tool for operational job management.
Analyses multiple weeks of planned job data to surface trends, team performance,
completion rates, and actionable insights across SPEN & ENW contracts.

Requirements:
pip install pandas openpyxl matplotlib

Usage:
python planned_jobs_analyser_v3.py

Author: Built for SPEN/ENW Operations
“””

import sys
import os
import re
import json
import glob
import threading
import webbrowser
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule

import matplotlib
matplotlib.use(‘TkAgg’)
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.ticker as mticker

# ═══════════════════════════════════════════════════════════════

# THEME & STYLING

# ═══════════════════════════════════════════════════════════════

DARK = {
‘bg’:         ‘#0B0F1A’,
‘card’:       ‘#111827’,
‘card_alt’:   ‘#1A2332’,
‘border’:     ‘#1E293B’,
‘fg’:         ‘#E2E8F0’,
‘fg_muted’:   ‘#94A3B8’,
‘fg_dim’:     ‘#64748B’,
‘accent’:     ‘#3B82F6’,
‘green’:      ‘#10B981’,
‘amber’:      ‘#F59E0B’,
‘red’:        ‘#EF4444’,
‘purple’:     ‘#8B5CF6’,
‘cyan’:       ‘#06B6D4’,
‘white’:      ‘#FFFFFF’,
‘entry_bg’:   ‘#1E293B’,
‘btn_primary’:’#3B82F6’,
‘btn_success’:’#10B981’,
‘btn_warning’:’#F59E0B’,
‘btn_danger’: ‘#EF4444’,
‘btn_purple’: ‘#8B5CF6’,
}

# Chart colour cycle

CHART_COLORS = [’#3B82F6’,’#10B981’,’#F59E0B’,’#EF4444’,’#8B5CF6’,’#06B6D4’,’#EC4899’,’#14B8A6’,’#F97316’,’#6366F1’]
STATUS_COLORS = {
‘Site Clear’:  ‘#10B981’,
‘In Progress’: ‘#3B82F6’,
‘Scheduled’:   ‘#F59E0B’,
‘On Hold’:     ‘#EF4444’,
‘Cancelled’:   ‘#64748B’,
}

# Excel styling constants

XL_DARK_HEADER   = PatternFill(‘solid’, fgColor=‘0B0F1A’)
XL_ACCENT_HEADER = PatternFill(‘solid’, fgColor=‘3B82F6’)
XL_GREEN_FILL    = PatternFill(‘solid’, fgColor=‘D1FAE5’)
XL_RED_FILL      = PatternFill(‘solid’, fgColor=‘FEE2E2’)
XL_AMBER_FILL    = PatternFill(‘solid’, fgColor=‘FEF3C7’)
XL_LIGHT_ROW     = PatternFill(‘solid’, fgColor=‘F1F5F9’)
XL_WHITE_ROW     = PatternFill(‘solid’, fgColor=‘FFFFFF’)
XL_PURPLE_FILL   = PatternFill(‘solid’, fgColor=‘EDE9FE’)

XL_HEADER_FONT   = Font(name=‘Aptos’, bold=True, color=‘FFFFFF’, size=11)
XL_TITLE_FONT    = Font(name=‘Aptos’, bold=True, color=‘0B0F1A’, size=16)
XL_SUBTITLE_FONT = Font(name=‘Aptos’, bold=True, color=‘3B82F6’, size=12)
XL_BODY_FONT     = Font(name=‘Aptos’, size=10, color=‘334155’)
XL_BOLD_FONT     = Font(name=‘Aptos’, bold=True, size=10, color=‘0F172A’)
XL_KPI_FONT      = Font(name=‘Aptos’, bold=True, size=28, color=‘0B0F1A’)
XL_KPI_LABEL     = Font(name=‘Aptos’, size=10, color=‘64748B’)

XL_BORDER = Border(
bottom=Side(style=‘thin’, color=‘E2E8F0’),
top=Side(style=‘thin’, color=‘E2E8F0’),
left=Side(style=‘thin’, color=‘E2E8F0’),
right=Side(style=‘thin’, color=‘E2E8F0’)
)
XL_THICK_BOTTOM = Border(bottom=Side(style=‘medium’, color=‘3B82F6’))

# ═══════════════════════════════════════════════════════════════

# SETTINGS MANAGER

# ═══════════════════════════════════════════════════════════════

class Settings:
def **init**(self):
self.path = Path.home() / ‘.jobs_analyser_settings.json’
self.defaults = {
‘auto_open’: True,
‘charts_in_excel’: True,
‘last_dir’: str(Path.home() / ‘Downloads’),
‘period_days’: 28,
‘window_geometry’: ‘1200x820’,
}
self.data = self._load()

```
def _load(self):
    try:
        if self.path.exists():
            with open(self.path) as f:
                d = json.load(f)
            for k, v in self.defaults.items():
                d.setdefault(k, v)
            return d
    except Exception:
        pass
    return dict(self.defaults)

def save(self):
    try:
        with open(self.path, 'w') as f:
            json.dump(self.data, f, indent=2)
    except Exception:
        pass

def get(self, key, default=None):
    return self.data.get(key, default)

def set(self, key, val):
    self.data[key] = val
    self.save()
```

# ═══════════════════════════════════════════════════════════════

# DATA LOADING & CLEANING

# ═══════════════════════════════════════════════════════════════

def extract_date_from_filename(filepath):
“”“Try to pull a date from the filename.”””
name = os.path.basename(filepath)
patterns = [
(r’(\d{4})-(\d{2})-(\d{2})’, ‘%Y-%m-%d’),
(r’(\d{2})-(\d{2})-(\d{4})’, ‘%d-%m-%Y’),
(r’(\d{4})(\d{2})(\d{2})’, ‘%Y%m%d’),
(r’(\d{2})*(\d{2})*(\d{4})’, ‘%d_%m_%Y’),
(r’(\d{4})*(\d{2})*(\d{2})’, ‘%Y_%m_%d’),
]
for pattern, fmt in patterns:
m = re.search(pattern, name)
if m:
try:
date_str = m.group(0)
return pd.to_datetime(date_str, format=fmt)
except Exception:
continue
return None

def load_file(filepath):
“”“Load an Excel/CSV file and return (DataFrame, date).”””
ext = os.path.splitext(filepath)[1].lower()
if ext == ‘.csv’:
df = pd.read_csv(filepath)
else:
df = pd.read_excel(filepath)

```
df.columns = df.columns.str.strip()

# Determine file date
file_date = extract_date_from_filename(filepath)
if file_date is None and 'Date From' in df.columns:
    try:
        file_date = pd.to_datetime(df['Date From'].dropna().iloc[0])
    except Exception:
        pass
if file_date is None:
    try:
        file_date = pd.Timestamp(datetime.fromtimestamp(os.path.getmtime(filepath)))
    except Exception:
        file_date = pd.Timestamp.now()

df['_analysis_date'] = file_date
df['_source_file'] = os.path.basename(filepath)
return df, file_date
```

# ═══════════════════════════════════════════════════════════════

# ANALYTICS ENGINE

# ═══════════════════════════════════════════════════════════════

class AnalyticsEngine:
“”“Crunches all the numbers from loaded data.”””

```
def __init__(self, all_data):
    """all_data: list of (DataFrame, Timestamp) tuples sorted by date."""
    self.all_data = sorted(all_data, key=lambda x: x[1])
    self.combined = pd.concat([df for df, _ in self.all_data], ignore_index=True)

@property
def date_range(self):
    return (self.all_data[0][1], self.all_data[-1][1])

@property
def total_files(self):
    return len(self.all_data)

@property
def total_jobs(self):
    return len(self.combined)

@property
def date_span_days(self):
    return (self.date_range[1] - self.date_range[0]).days + 1

def status_summary(self):
    return self.combined['Job Status'].value_counts().to_dict()

def completion_rate(self):
    ss = self.status_summary()
    total = sum(ss.values())
    return (ss.get('Site Clear', 0) / total * 100) if total > 0 else 0

def daily_trends(self):
    """Per-file date breakdown."""
    rows = []
    for df, dt in self.all_data:
        sc = df['Job Status'].value_counts()
        rows.append({
            'date': dt,
            'label': dt.strftime('%d %b'),
            'total': len(df),
            'completed': sc.get('Site Clear', 0),
            'in_progress': sc.get('In Progress', 0),
            'scheduled': sc.get('Scheduled', 0),
            'on_hold': sc.get('On Hold', 0),
            'cancelled': sc.get('Cancelled', 0),
            'completion_rate': (sc.get('Site Clear', 0) / len(df) * 100) if len(df) > 0 else 0,
        })
    return rows

def team_performance(self):
    """Per-team aggregated metrics."""
    if 'Gang Ref' not in self.combined.columns:
        return []
    teams = {}
    for _, row in self.combined.iterrows():
        t = row.get('Gang Ref', 'Unassigned')
        if pd.isna(t):
            t = 'Unassigned'
        if t not in teams:
            teams[t] = {'name': t, 'total': 0, 'completed': 0, 'in_progress': 0, 'weekly': defaultdict(int)}
        teams[t]['total'] += 1
        if row.get('Job Status') == 'Site Clear':
            teams[t]['completed'] += 1
        if row.get('Job Status') == 'In Progress':
            teams[t]['in_progress'] += 1

    # Weekly sparkline data
    for i, (df, dt) in enumerate(self.all_data):
        if 'Gang Ref' in df.columns:
            for _, row in df.iterrows():
                t = row.get('Gang Ref', 'Unassigned')
                if pd.isna(t):
                    t = 'Unassigned'
                if t in teams:
                    teams[t]['weekly'][i] += 1

    result = []
    for t in teams.values():
        t['completion_rate'] = (t['completed'] / t['total'] * 100) if t['total'] > 0 else 0
        t['sparkline'] = [t['weekly'].get(i, 0) for i in range(len(self.all_data))]
        result.append(t)
    return sorted(result, key=lambda x: x['total'], reverse=True)

def pm_performance(self):
    """Per-PM aggregated metrics."""
    col = 'Project Manager'
    if col not in self.combined.columns:
        return []
    pms = {}
    for _, row in self.combined.iterrows():
        p = row.get(col, 'Unassigned')
        if pd.isna(p):
            p = 'Unassigned'
        if p not in pms:
            pms[p] = {'name': p, 'total': 0, 'completed': 0}
        pms[p]['total'] += 1
        if row.get('Job Status') == 'Site Clear':
            pms[p]['completed'] += 1
    for p in pms.values():
        p['completion_rate'] = (p['completed'] / p['total'] * 100) if p['total'] > 0 else 0
    return sorted(pms.values(), key=lambda x: x['total'], reverse=True)

def work_type_breakdown(self):
    if 'Work Type' not in self.combined.columns:
        return {}
    return self.combined['Work Type'].value_counts().to_dict()

def postcode_breakdown(self):
    if 'Postcode' not in self.combined.columns:
        return {}
    pc = self.combined['Postcode'].dropna().str[:3]
    return pc.value_counts().head(15).to_dict()

def contract_breakdown(self):
    if 'Contract' not in self.combined.columns:
        return {}
    return self.combined['Contract'].value_counts().to_dict()

def week_over_week_deltas(self):
    trends = self.daily_trends()
    deltas = []
    for i, t in enumerate(trends):
        d = dict(t)
        if i > 0:
            prev = trends[i - 1]
            d['delta_total'] = t['total'] - prev['total']
            d['delta_rate'] = round(t['completion_rate'] - prev['completion_rate'], 1)
        else:
            d['delta_total'] = 0
            d['delta_rate'] = 0.0
        deltas.append(d)
    return deltas
```

# ═══════════════════════════════════════════════════════════════

# EXCEL REPORT GENERATOR

# ═══════════════════════════════════════════════════════════════

def _xl_header(ws, row, headers):
for c, h in enumerate(headers, 1):
cell = ws.cell(row=row, column=c, value=h)
cell.fill = XL_DARK_HEADER
cell.font = XL_HEADER_FONT
cell.alignment = Alignment(horizontal=‘center’, vertical=‘center’, wrap_text=True)
cell.border = XL_BORDER

def _xl_data_rows(ws, start, data_rows, headers_count):
for i, row_data in enumerate(data_rows):
fill = XL_LIGHT_ROW if i % 2 == 0 else XL_WHITE_ROW
for c, val in enumerate(row_data, 1):
cell = ws.cell(row=start + i, column=c, value=val)
cell.font = XL_BODY_FONT
cell.fill = fill
cell.border = XL_BORDER
cell.alignment = Alignment(vertical=‘center’, wrap_text=True)

def _xl_auto_width(ws, max_col, max_w=45):
for c in range(1, max_col + 1):
mx = 0
for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
for cell in row:
if cell.value:
mx = max(mx, len(str(cell.value)))
ws.column_dimensions[get_column_letter(c)].width = min(mx + 4, max_w)

def _xl_write_table(ws, row, headers, data):
_xl_header(ws, row, headers)
if data:
_xl_data_rows(ws, row + 1, data, len(headers))
_xl_auto_width(ws, len(headers))
return row + len(data) + 2

def generate_excel_report(engine, output_dir, settings=None):
“”“Generate the full multi-sheet Excel report.”””
wb = Workbook()

```
# ─── SHEET 1: EXECUTIVE DASHBOARD ───
ws = wb.active
ws.title = 'Executive Dashboard'
ws.sheet_properties.tabColor = '3B82F6'
ws.sheet_view.showGridLines = False

r = 1
ws.cell(row=r, column=1, value='MULTI-WEEK OPERATIONS REPORT').font = XL_TITLE_FONT
r += 1
dr = engine.date_range
ws.cell(row=r, column=1, value=f'{dr[0].strftime("%d %b %Y")} — {dr[1].strftime("%d %b %Y")}  ·  {engine.total_files} reporting periods  ·  Generated {datetime.now().strftime("%d %b %Y %H:%M")}').font = Font(name='Aptos', size=10, color='64748B')
r += 2

# KPI Row
ss = engine.status_summary()
kpis = [
    ('Total Jobs', engine.total_jobs),
    ('Site Clear', ss.get('Site Clear', 0)),
    ('In Progress', ss.get('In Progress', 0)),
    ('Completion Rate', f'{engine.completion_rate():.1f}%'),
    ('Active Teams', len(engine.team_performance())),
    ('Reporting Period', f'{engine.date_span_days} days'),
]
for i, (label, val) in enumerate(kpis):
    col = 1 + i * 2
    ws.cell(row=r, column=col, value=val).font = XL_KPI_FONT
    ws.cell(row=r + 1, column=col, value=label).font = XL_KPI_LABEL
r += 4

# Trend table
ws.cell(row=r, column=1, value='WEEKLY TRENDS').font = XL_SUBTITLE_FONT
r += 1
trends = engine.daily_trends()
headers = ['Date', 'Total', 'Completed', 'In Progress', 'Scheduled', 'On Hold', 'Completion %']
trend_data = [(t['label'], t['total'], t['completed'], t['in_progress'], t['scheduled'],
               t['on_hold'], f"{t['completion_rate']:.1f}%") for t in trends]
trend_start = r
r = _xl_write_table(ws, r, headers, trend_data)

# Conditional formatting on completion rate column
for i, t in enumerate(trends):
    cell = ws.cell(row=trend_start + 1 + i, column=7)
    if t['completion_rate'] >= 40:
        cell.fill = XL_GREEN_FILL
        cell.font = Font(name='Aptos', bold=True, size=10, color='065F46')
    elif t['completion_rate'] >= 25:
        cell.fill = XL_AMBER_FILL
        cell.font = Font(name='Aptos', bold=True, size=10, color='92400E')
    else:
        cell.fill = XL_RED_FILL
        cell.font = Font(name='Aptos', bold=True, size=10, color='991B1B')

# Line chart: trends over time
try:
    chart = LineChart()
    chart.title = 'Job Volume & Completion Trend'
    chart.height = 14
    chart.width = 24
    chart.y_axis.title = 'Count'

    data_ref = Reference(ws, min_col=2, min_row=trend_start, max_col=5, max_row=trend_start + len(trends))
    cats = Reference(ws, min_col=1, min_row=trend_start + 1, max_row=trend_start + len(trends))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    for i, color in enumerate(['3B82F6', '10B981', 'F59E0B', 'EF4444']):
        if i < len(chart.series):
            chart.series[i].graphicalProperties.line.solidFill = color

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = False
    ws.add_chart(chart, f'A{r}')
    r += 16
except Exception:
    r += 1

# ─── SHEET 2: TEAM PERFORMANCE ───
ws2 = wb.create_sheet('Team Performance')
ws2.sheet_properties.tabColor = '10B981'
ws2.sheet_view.showGridLines = False

r2 = 1
ws2.cell(row=r2, column=1, value='TEAM PERFORMANCE ANALYSIS').font = XL_TITLE_FONT
r2 += 2

team_data = engine.team_performance()
if team_data:
    headers = ['Team', 'Total Jobs', 'Completed', 'In Progress', 'Completion %', 'Rank']
    rows = []
    for i, t in enumerate(sorted(team_data, key=lambda x: x['completion_rate'], reverse=True)):
        rows.append((t['name'], t['total'], t['completed'], t['in_progress'],
                    f"{t['completion_rate']:.1f}%", i + 1))
    team_start = r2
    r2 = _xl_write_table(ws2, r2, headers, rows)

    # Colour the completion rate and rank
    for i, t in enumerate(sorted(team_data, key=lambda x: x['completion_rate'], reverse=True)):
        rate_cell = ws2.cell(row=team_start + 1 + i, column=5)
        rank_cell = ws2.cell(row=team_start + 1 + i, column=6)
        if t['completion_rate'] >= 40:
            rate_cell.fill = XL_GREEN_FILL
            rate_cell.font = Font(name='Aptos', bold=True, size=10, color='065F46')
        elif t['completion_rate'] >= 25:
            rate_cell.fill = XL_AMBER_FILL
        else:
            rate_cell.fill = XL_RED_FILL

        if i == 0:
            rank_cell.fill = PatternFill('solid', fgColor='FEF08A')
            rank_cell.font = Font(name='Aptos', bold=True, size=10, color='854D0E')

    # Team bar chart
    try:
        chart2 = BarChart()
        chart2.type = 'col'
        chart2.title = 'Team Job Volumes'
        chart2.height = 14
        chart2.width = 22

        data_ref = Reference(ws2, min_col=2, min_row=team_start, max_col=4, max_row=team_start + len(rows))
        cats = Reference(ws2, min_col=1, min_row=team_start + 1, max_row=team_start + len(rows))
        chart2.add_data(data_ref, titles_from_data=True)
        chart2.set_categories(cats)

        for i, color in enumerate(['3B82F6', '10B981', 'F59E0B']):
            if i < len(chart2.series):
                chart2.series[i].graphicalProperties.solidFill = color

        ws2.add_chart(chart2, f'A{r2}')
        r2 += 16
    except Exception:
        pass

    # Team weekly breakdown
    ws2.cell(row=r2, column=1, value='TEAM WEEKLY BREAKDOWN').font = XL_SUBTITLE_FONT
    r2 += 1
    all_labels = [t['label'] for t in trends]
    tw_headers = ['Team'] + all_labels + ['Total', 'Avg/Week']
    tw_rows = []
    for t in team_data:
        row_vals = [t['name']] + t['sparkline']
        row_vals.append(t['total'])
        num_weeks = max(len(t['sparkline']), 1)
        row_vals.append(f"{t['total'] / num_weeks:.1f}")
        tw_rows.append(row_vals)
    _xl_write_table(ws2, r2, tw_headers, tw_rows)

# ─── SHEET 3: PM LEADERBOARD ───
ws3 = wb.create_sheet('PM Leaderboard')
ws3.sheet_properties.tabColor = '8B5CF6'
ws3.sheet_view.showGridLines = False

r3 = 1
ws3.cell(row=r3, column=1, value='PROJECT MANAGER LEADERBOARD').font = XL_TITLE_FONT
r3 += 2

pm_data = engine.pm_performance()
if pm_data:
    pm_sorted = sorted(pm_data, key=lambda x: x['completion_rate'], reverse=True)
    headers = ['Rank', 'Project Manager', 'Total Jobs', 'Completed', 'Completion %']
    rows = [(i + 1, p['name'], p['total'], p['completed'], f"{p['completion_rate']:.1f}%")
            for i, p in enumerate(pm_sorted)]
    pm_start = r3
    r3 = _xl_write_table(ws3, r3, headers, rows)

    for i, p in enumerate(pm_sorted):
        cell = ws3.cell(row=pm_start + 1 + i, column=5)
        if p['completion_rate'] >= 40:
            cell.fill = XL_GREEN_FILL
        elif p['completion_rate'] >= 25:
            cell.fill = XL_AMBER_FILL
        else:
            cell.fill = XL_RED_FILL

    try:
        chart3 = BarChart()
        chart3.type = 'bar'
        chart3.title = 'PM Completion Rates'
        chart3.height = 12
        chart3.width = 20

        # We need numeric values for chart, so write them separately
        pm_chart_start = r3 + 1
        ws3.cell(row=pm_chart_start, column=1, value='PM').font = XL_HEADER_FONT
        ws3.cell(row=pm_chart_start, column=2, value='Rate %').font = XL_HEADER_FONT
        for i, p in enumerate(pm_sorted):
            ws3.cell(row=pm_chart_start + 1 + i, column=1, value=p['name'])
            ws3.cell(row=pm_chart_start + 1 + i, column=2, value=round(p['completion_rate'], 1))

        data_ref = Reference(ws3, min_col=2, min_row=pm_chart_start, max_row=pm_chart_start + len(pm_sorted))
        cats = Reference(ws3, min_col=1, min_row=pm_chart_start + 1, max_row=pm_chart_start + len(pm_sorted))
        chart3.add_data(data_ref, titles_from_data=True)
        chart3.set_categories(cats)
        if chart3.series:
            chart3.series[0].graphicalProperties.solidFill = '8B5CF6'
        ws3.add_chart(chart3, f'D{pm_start}')
    except Exception:
        pass

# ─── SHEET 4: WEEK-OVER-WEEK ───
ws4 = wb.create_sheet('Week-over-Week')
ws4.sheet_properties.tabColor = 'F59E0B'
ws4.sheet_view.showGridLines = False

r4 = 1
ws4.cell(row=r4, column=1, value='WEEK-OVER-WEEK COMPARISON').font = XL_TITLE_FONT
r4 += 2

wow = engine.week_over_week_deltas()
if wow:
    headers = ['Date', 'Total', 'Completed', 'In Progress', 'Scheduled', 'On Hold',
               'Completion %', 'Δ Total', 'Δ Rate (pp)']
    rows = []
    for w in wow:
        delta_t = w['delta_total']
        delta_r = w['delta_rate']
        dt_str = f"+{delta_t}" if delta_t > 0 else str(delta_t) if delta_t != 0 else "—"
        dr_str = f"+{delta_r}" if delta_r > 0 else str(delta_r) if delta_r != 0 else "—"
        rows.append((w['label'], w['total'], w['completed'], w['in_progress'], w['scheduled'],
                    w['on_hold'], f"{w['completion_rate']:.1f}%", dt_str, dr_str))
    wow_start = r4
    r4 = _xl_write_table(ws4, r4, headers, rows)

    # Colour delta columns
    for i, w in enumerate(wow):
        dt_cell = ws4.cell(row=wow_start + 1 + i, column=8)
        dr_cell = ws4.cell(row=wow_start + 1 + i, column=9)
        if w['delta_total'] > 0:
            dt_cell.font = Font(name='Aptos', bold=True, size=10, color='065F46')
        elif w['delta_total'] < 0:
            dt_cell.font = Font(name='Aptos', bold=True, size=10, color='991B1B')
        if w['delta_rate'] > 0:
            dr_cell.font = Font(name='Aptos', bold=True, size=10, color='065F46')
        elif w['delta_rate'] < 0:
            dr_cell.font = Font(name='Aptos', bold=True, size=10, color='991B1B')

# ─── SHEET 5: ALL JOBS ───
ws5 = wb.create_sheet('All Jobs')
ws5.sheet_properties.tabColor = '06B6D4'
ws5.sheet_view.showGridLines = False

r5 = 1
ws5.cell(row=r5, column=1, value='CONSOLIDATED JOB LIST').font = XL_TITLE_FONT
r5 += 2

cols = ['_analysis_date', 'Job ID', 'Job Status', 'Gang Ref', 'Work Type',
        'Project Manager', 'Postcode', 'Contract']
cols_available = [c for c in cols if c in engine.combined.columns]

display_names = {'_analysis_date': 'Report Date'}
_xl_header(ws5, r5, [display_names.get(c, c) for c in cols_available])

for i, (_, row) in enumerate(engine.combined[cols_available].iterrows()):
    for c_idx, col_name in enumerate(cols_available, 1):
        val = row[col_name]
        if hasattr(val, 'strftime'):
            val = val.strftime('%Y-%m-%d')
        cell = ws5.cell(row=r5 + 1 + i, column=c_idx, value=val)
        cell.font = XL_BODY_FONT
        cell.border = XL_BORDER

        # Status colour coding
        status = row.get('Job Status', '')
        if c_idx == 1:  # Apply fill to whole row based on status
            pass
    # Row fill by status
    status = row.get('Job Status', '')
    if status == 'Site Clear':
        fill = XL_GREEN_FILL
    elif status == 'In Progress':
        fill = PatternFill('solid', fgColor='DBEAFE')
    elif status == 'Scheduled':
        fill = XL_AMBER_FILL
    elif status == 'On Hold':
        fill = XL_RED_FILL
    else:
        fill = XL_WHITE_ROW if i % 2 else XL_LIGHT_ROW
    for c_idx in range(1, len(cols_available) + 1):
        ws5.cell(row=r5 + 1 + i, column=c_idx).fill = fill

_xl_auto_width(ws5, len(cols_available))
last_row = r5 + len(engine.combined)
ws5.auto_filter.ref = f'A{r5}:{get_column_letter(len(cols_available))}{last_row}'
ws5.freeze_panes = f'A{r5 + 1}'

# ─── SAVE ───
start_str = dr[0].strftime('%Y%m%d')
end_str = dr[1].strftime('%Y%m%d')
filename = f'multi_week_analysis_{start_str}_to_{end_str}.xlsx'
output_path = os.path.join(output_dir, filename)
wb.save(output_path)
return output_path
```

# ═══════════════════════════════════════════════════════════════

# MATPLOTLIB DASHBOARD CHARTS

# ═══════════════════════════════════════════════════════════════

def configure_mpl_dark():
“”“Configure matplotlib for dark theme charts.”””
plt.rcParams.update({
‘figure.facecolor’: DARK[‘card’],
‘axes.facecolor’: DARK[‘bg’],
‘axes.edgecolor’: DARK[‘border’],
‘axes.labelcolor’: DARK[‘fg_muted’],
‘text.color’: DARK[‘fg’],
‘xtick.color’: DARK[‘fg_dim’],
‘ytick.color’: DARK[‘fg_dim’],
‘grid.color’: DARK[‘border’],
‘grid.alpha’: 0.5,
‘font.family’: ‘sans-serif’,
‘font.size’: 9,
})

def create_overview_figure(engine, width_px=1100, height_px=500):
“”“Create the overview dashboard figure with 4 subplots.”””
configure_mpl_dark()
dpi = 100
fig = Figure(figsize=(width_px / dpi, height_px / dpi), dpi=dpi)
fig.patch.set_facecolor(DARK[‘card’])
fig.subplots_adjust(hspace=0.45, wspace=0.35, left=0.06, right=0.96, top=0.92, bottom=0.1)

```
trends = engine.daily_trends()
labels = [t['label'] for t in trends]

# 1. Stacked area — job pipeline
ax1 = fig.add_subplot(2, 2, 1)
completed = [t['completed'] for t in trends]
in_prog = [t['in_progress'] for t in trends]
scheduled = [t['scheduled'] for t in trends]
ax1.stackplot(labels, completed, in_prog, scheduled,
              colors=[DARK['green'], DARK['accent'], DARK['amber']], alpha=0.8,
              labels=['Completed', 'In Progress', 'Scheduled'])
ax1.set_title('Job Pipeline', color=DARK['white'], fontsize=11, fontweight='bold', pad=10)
ax1.legend(loc='upper left', fontsize=7, framealpha=0.3)
ax1.tick_params(axis='x', rotation=30)

# 2. Completion rate line
ax2 = fig.add_subplot(2, 2, 2)
rates = [t['completion_rate'] for t in trends]
ax2.plot(labels, rates, color=DARK['green'], linewidth=2.5, marker='o', markersize=6,
         markerfacecolor=DARK['white'], markeredgecolor=DARK['green'], markeredgewidth=2)
ax2.fill_between(labels, rates, alpha=0.15, color=DARK['green'])
ax2.set_title('Completion Rate %', color=DARK['white'], fontsize=11, fontweight='bold', pad=10)
ax2.set_ylim(0, max(max(rates) * 1.2, 10))
ax2.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.0f%%'))
ax2.tick_params(axis='x', rotation=30)

# 3. Status pie
ax3 = fig.add_subplot(2, 2, 3)
ss = engine.status_summary()
if ss:
    pie_labels = list(ss.keys())
    pie_vals = list(ss.values())
    pie_colors = [STATUS_COLORS.get(s, DARK['fg_dim']) for s in pie_labels]
    wedges, texts, autotexts = ax3.pie(pie_vals, labels=pie_labels, colors=pie_colors,
                                        autopct='%1.0f%%', startangle=90, pctdistance=0.8,
                                        textprops={'fontsize': 8, 'color': DARK['fg']})
    for at in autotexts:
        at.set_fontsize(7)
        at.set_color(DARK['white'])
    ax3.set_title('Status Split', color=DARK['white'], fontsize=11, fontweight='bold', pad=10)

# 4. Team bar chart
ax4 = fig.add_subplot(2, 2, 4)
team_data = engine.team_performance()[:8]
if team_data:
    t_names = [t['name'] for t in team_data]
    t_completed = [t['completed'] for t in team_data]
    t_other = [t['total'] - t['completed'] for t in team_data]
    y_pos = range(len(t_names))
    ax4.barh(y_pos, t_completed, color=DARK['green'], alpha=0.9, label='Completed', height=0.6)
    ax4.barh(y_pos, t_other, left=t_completed, color=DARK['accent'], alpha=0.5, label='Other', height=0.6)
    ax4.set_yticks(y_pos)
    ax4.set_yticklabels(t_names, fontsize=8)
    ax4.set_title('Team Volumes', color=DARK['white'], fontsize=11, fontweight='bold', pad=10)
    ax4.legend(loc='lower right', fontsize=7, framealpha=0.3)
    ax4.invert_yaxis()

return fig
```

# ═══════════════════════════════════════════════════════════════

# MAIN APPLICATION

# ═══════════════════════════════════════════════════════════════

class MultiWeekAnalyserApp:
def **init**(self):
self.settings = Settings()
self.root = tk.Tk()
self.root.title(‘Multi-Week Jobs Analyser v3.0’)

```
    geom = self.settings.get('window_geometry', '1200x820')
    self.root.geometry(geom)
    self.root.minsize(900, 650)
    self.root.configure(bg=DARK['bg'])

    # Try to set icon (won't fail if not available)
    try:
        self.root.iconname('Jobs Analyser')
    except Exception:
        pass

    self.selected_files = []
    self.engine = None
    self.chart_canvas = None

    self._build_ui()

# ─── UI CONSTRUCTION ───

def _build_ui(self):
    bg = DARK['bg']

    # Top bar
    top = tk.Frame(self.root, bg=DARK['card'], height=60)
    top.pack(fill='x')
    top.pack_propagate(False)

    title_frame = tk.Frame(top, bg=DARK['card'])
    title_frame.pack(side='left', padx=20, pady=10)
    tk.Label(title_frame, text='⚡ Multi-Week Jobs Analyser',
             font=('Segoe UI', 18, 'bold'), bg=DARK['card'], fg=DARK['white']).pack(anchor='w')
    self.subtitle_var = tk.StringVar(value='Load files to begin analysis')
    tk.Label(title_frame, textvariable=self.subtitle_var,
             font=('Segoe UI', 10), bg=DARK['card'], fg=DARK['fg_dim']).pack(anchor='w')

    # Buttons in top bar
    btn_frame = tk.Frame(top, bg=DARK['card'])
    btn_frame.pack(side='right', padx=20, pady=10)

    self._make_btn(btn_frame, '📁 Add Files', self.add_files, DARK['accent']).pack(side='left', padx=3)
    self._make_btn(btn_frame, '🗂 Add Folder', self.add_folder, DARK['amber']).pack(side='left', padx=3)
    self._make_btn(btn_frame, '🔍 Auto-Detect', self.auto_detect, DARK['purple']).pack(side='left', padx=3)
    self._make_btn(btn_frame, '🗑 Clear', self.clear_files, DARK['red']).pack(side='left', padx=3)

    # Main content — split into left panel (file list) and right panel (charts/results)
    main = tk.PanedWindow(self.root, orient='horizontal', bg=bg, sashwidth=2, sashrelief='flat')
    main.pack(fill='both', expand=True, padx=0, pady=0)

    # Left panel
    left = tk.Frame(main, bg=bg, width=320)
    main.add(left, minsize=260)

    # File list
    file_label_frame = tk.Frame(left, bg=bg)
    file_label_frame.pack(fill='x', padx=12, pady=(12, 4))
    tk.Label(file_label_frame, text='SELECTED FILES', font=('Segoe UI', 9, 'bold'),
             bg=bg, fg=DARK['fg_dim']).pack(side='left')
    self.file_count_var = tk.StringVar(value='0')
    tk.Label(file_label_frame, textvariable=self.file_count_var,
             font=('Segoe UI', 9), bg=bg, fg=DARK['accent']).pack(side='right')

    list_frame = tk.Frame(left, bg=DARK['card'], bd=1, relief='solid', highlightbackground=DARK['border'])
    list_frame.pack(fill='both', expand=True, padx=12, pady=4)

    scrollbar = tk.Scrollbar(list_frame, bg=DARK['border'])
    scrollbar.pack(side='right', fill='y')

    self.file_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set,
                                    font=('Consolas', 9), height=15,
                                    bg=DARK['card'], fg=DARK['fg'],
                                    selectbackground=DARK['accent'],
                                    selectforeground=DARK['white'],
                                    bd=0, highlightthickness=0,
                                    activestyle='none')
    self.file_listbox.pack(side='left', fill='both', expand=True)
    scrollbar.config(command=self.file_listbox.yview)

    # Analyse button
    self.analyse_btn = tk.Button(left, text='🚀  ANALYSE TRENDS', command=self._run,
                                 font=('Segoe UI', 13, 'bold'),
                                 bg=DARK['btn_success'], fg=DARK['white'],
                                 activebackground='#059669', activeforeground=DARK['white'],
                                 bd=0, padx=20, pady=12, cursor='hand2',
                                 state='disabled', relief='flat')
    self.analyse_btn.pack(fill='x', padx=12, pady=12)

    # Status
    status_frame = tk.Frame(left, bg=bg)
    status_frame.pack(fill='x', padx=12, pady=(0, 8))

    self.status_var = tk.StringVar(value='Ready')
    self.status_label = tk.Label(status_frame, textvariable=self.status_var,
                                 font=('Segoe UI', 9), bg=bg, fg=DARK['fg_muted'],
                                 wraplength=280, justify='left', anchor='w')
    self.status_label.pack(fill='x')

    self.progress = ttk.Progressbar(status_frame, mode='indeterminate', length=280)
    self.progress.pack(fill='x', pady=(4, 0))

    # Right panel (charts)
    self.right_panel = tk.Frame(main, bg=bg)
    main.add(self.right_panel, minsize=500)

    # Placeholder
    self.placeholder = tk.Frame(self.right_panel, bg=bg)
    self.placeholder.pack(fill='both', expand=True)

    ph_inner = tk.Frame(self.placeholder, bg=bg)
    ph_inner.place(relx=0.5, rely=0.5, anchor='center')

    tk.Label(ph_inner, text='📊', font=('Segoe UI', 48), bg=bg, fg=DARK['fg_dim']).pack()
    tk.Label(ph_inner, text='Add files and run analysis\nto see interactive charts here',
             font=('Segoe UI', 14), bg=bg, fg=DARK['fg_dim'], justify='center').pack(pady=10)

def _make_btn(self, parent, text, cmd, color):
    return tk.Button(parent, text=text, command=cmd,
                    font=('Segoe UI', 9, 'bold'), bg=color, fg=DARK['white'],
                    activebackground=color, activeforeground=DARK['white'],
                    bd=0, padx=12, pady=6, cursor='hand2', relief='flat')

# ─── FILE MANAGEMENT ───

def add_files(self):
    files = filedialog.askopenfilenames(
        title='Select Excel/CSV files',
        initialdir=self.settings.get('last_dir'),
        filetypes=[('Excel files', '*.xlsx *.xls'), ('CSV files', '*.csv'), ('All', '*.*')]
    )
    added = 0
    for f in files:
        if f not in self.selected_files:
            self.selected_files.append(f)
            added += 1
    if files:
        self.settings.set('last_dir', os.path.dirname(files[0]))
    self._refresh_list()
    if added:
        self.status_var.set(f'✅ Added {added} file(s)')

def add_folder(self):
    folder = filedialog.askdirectory(
        title='Select folder with Excel/CSV files',
        initialdir=self.settings.get('last_dir')
    )
    if not folder:
        return
    added = 0
    for ext in ('*.xlsx', '*.xls', '*.csv'):
        for f in glob.glob(os.path.join(folder, ext)):
            if f not in self.selected_files:
                self.selected_files.append(f)
                added += 1
    self.settings.set('last_dir', folder)
    self._refresh_list()
    self.status_var.set(f'✅ Added {added} file(s) from folder' if added else '⚠️ No new files found')

def auto_detect(self):
    days = self.settings.get('period_days', 28)
    cutoff = datetime.now() - timedelta(days=days)
    search_dirs = [
        Path.home() / 'Downloads',
        Path.home() / 'Desktop',
        Path(self.settings.get('last_dir', '')),
    ]
    keywords = ['job', 'planned', 'export', 'report', 'week', 'spen', 'enw']
    added = 0
    for d in search_dirs:
        if not d.exists():
            continue
        for ext in ('*.xlsx', '*.xls', '*.csv'):
            for f in d.glob(ext):
                try:
                    if f.stat().st_mtime < cutoff.timestamp():
                        continue
                    if any(kw in f.name.lower() for kw in keywords):
                        fp = str(f)
                        if fp not in self.selected_files:
                            self.selected_files.append(fp)
                            added += 1
                except Exception:
                    continue
    self._refresh_list()
    self.status_var.set(f'✅ Auto-detected {added} file(s) from last {days} days' if added else f'⚠️ No matching files in last {days} days')

def clear_files(self):
    self.selected_files = []
    self._refresh_list()
    self.engine = None
    self._show_placeholder()
    self.status_var.set('🗑 Cleared')

def _refresh_list(self):
    self.file_listbox.delete(0, tk.END)
    try:
        sorted_files = sorted(self.selected_files, key=lambda x: os.path.getmtime(x), reverse=True)
    except Exception:
        sorted_files = self.selected_files

    for f in sorted_files:
        name = os.path.basename(f)
        try:
            mt = datetime.fromtimestamp(os.path.getmtime(f)).strftime('%d %b')
            display = f'{mt}  ·  {name}'
        except Exception:
            display = name
        self.file_listbox.insert(tk.END, display)

    n = len(self.selected_files)
    self.file_count_var.set(f'{n} file{"s" if n != 1 else ""}')
    self.analyse_btn.config(state='normal' if n > 0 else 'disabled')

# ─── ANALYSIS ───

def _run(self):
    if not self.selected_files:
        return
    self.analyse_btn.config(state='disabled')
    self.progress.start(12)
    self.status_var.set('🔄 Loading and processing files...')

    def work():
        try:
            all_data = []
            failed = []
            total = len(self.selected_files)

            for i, fp in enumerate(self.selected_files):
                self.root.after(0, lambda i=i: self.status_var.set(
                    f'🔄 Processing file {i + 1}/{total}...'))
                try:
                    df, dt = load_file(fp)
                    all_data.append((df, dt))
                except Exception as e:
                    failed.append((fp, str(e)))

            if not all_data:
                raise Exception('No files loaded successfully')

            self.root.after(0, lambda: self.status_var.set('🔄 Crunching analytics...'))
            engine = AnalyticsEngine(all_data)

            self.root.after(0, lambda: self.status_var.set('🔄 Generating Excel report...'))
            output_dir = os.path.dirname(os.path.abspath(self.selected_files[0]))
            output_path = generate_excel_report(engine, output_dir, self.settings)

            self.root.after(0, lambda: self._done(engine, output_path, failed))
        except Exception as e:
            self.root.after(0, lambda msg=str(e): self._error(msg))

    threading.Thread(target=work, daemon=True).start()

def _done(self, engine, output_path, failed):
    self.progress.stop()
    self.engine = engine
    self.analyse_btn.config(state='normal')

    dr = engine.date_range
    days = engine.date_span_days
    fail_msg = f'  ⚠️ {len(failed)} failed' if failed else ''

    self.status_var.set(
        f'✅ Done! {engine.total_files} files · {engine.total_jobs} jobs · '
        f'{days} days · {engine.completion_rate():.1f}% completion{fail_msg}'
    )
    self.subtitle_var.set(
        f'{dr[0].strftime("%d %b %Y")} — {dr[1].strftime("%d %b %Y")}  ·  '
        f'{engine.total_files} files  ·  {engine.total_jobs:,} total jobs'
    )

    # Show charts
    self._show_charts(engine)

    # Ask to open
    msg = (f'Analysis Complete!\n\n'
           f'📅 Period: {dr[0].strftime("%d %b %Y")} — {dr[1].strftime("%d %b %Y")} ({days} days)\n'
           f'📁 Files: {engine.total_files}\n'
           f'📊 Jobs: {engine.total_jobs:,}\n'
           f'✅ Completion: {engine.completion_rate():.1f}%\n'
           f'👥 Teams: {len(engine.team_performance())}\n'
           f'{fail_msg}\n\n'
           f'Saved: {os.path.basename(output_path)}\n\n'
           f'Open the Excel report now?')

    if messagebox.askyesno('Analysis Complete', msg):
        try:
            os.startfile(output_path)
        except AttributeError:
            webbrowser.open(f'file://{output_path}')

def _error(self, msg):
    self.progress.stop()
    self.analyse_btn.config(state='normal')
    self.status_var.set(f'❌ {msg}')
    messagebox.showerror('Error', f'Analysis failed:\n\n{msg}')

# ─── CHART DISPLAY ───

def _show_placeholder(self):
    if self.chart_canvas:
        self.chart_canvas.get_tk_widget().destroy()
        self.chart_canvas = None
    self.placeholder.pack(fill='both', expand=True)

def _show_charts(self, engine):
    self.placeholder.pack_forget()

    if self.chart_canvas:
        self.chart_canvas.get_tk_widget().destroy()

    # Get actual pixel dimensions
    self.right_panel.update_idletasks()
    w = max(self.right_panel.winfo_width(), 600)
    h = max(self.right_panel.winfo_height(), 400)

    fig = create_overview_figure(engine, width_px=w, height_px=h)
    self.chart_canvas = FigureCanvasTkAgg(fig, master=self.right_panel)
    self.chart_canvas.draw()
    self.chart_canvas.get_tk_widget().pack(fill='both', expand=True)

# ─── RUN ───

def run(self):
    self.root.mainloop()
```

# ═══════════════════════════════════════════════════════════════

# ENTRY POINT

# ═══════════════════════════════════════════════════════════════

if **name** == ‘**main**’:
app = MultiWeekAnalyserApp()
app.run()
